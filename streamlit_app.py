import streamlit as st
import pandas as pd
import pdfplumber
import unidecode
from rapidfuzz import fuzz
import jinja2
import datetime
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def normalize(s):
    if pd.isna(s) or s is None:
        return ""
    return unidecode.unidecode(str(s)).replace(" ", "").lower()

def normalize_order_number(order_number):
    """Normalise strictement un numéro de commande pour le matching.
    
    Supprime tous les espaces et caractères spéciaux, convertit en string,
    normalisation stricte pour uniformiser l'extraction et la comparaison.
    """
    if pd.isna(order_number) or order_number is None:
        return ""
    
    # Convertir en string
    order_str = str(order_number).strip()
    
    # Supprimer le .0 d'Excel si présent (ex: 5600025054.0 -> 5600025054)
    if order_str.endswith('.0'):
        order_str = order_str[:-2]
    
    # Supprimer tous les espaces et caractères spéciaux, ne garder que les chiffres
    order_clean = re.sub(r'[^0-9]', '', order_str)
    
    # Valider que c'est un numéro valide (au moins 8 chiffres)
    if len(order_clean) >= 8:
        return order_clean
    
    return ""

def extract_pdf_lines(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            lines = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines.extend(text.split("\n"))
            return lines
    except Exception:
        return []

def extract_commande_from_pdf(pdf_lines):
    """Extrait le numéro de commande depuis les lignes du PDF avec normalisation stricte"""
    found_orders = []
    
    for line in pdf_lines:
        # Chercher tous les patterns possibles de numéros de commande
        # Pattern principal: 56 suivi de 8 chiffres (total 10 chiffres)
        matches = re.findall(r'56\d{8}', line)
        found_orders.extend(matches)
        
        # Pattern alternatif: 560 suivi de 7 chiffres (total 10 chiffres)
        matches = re.findall(r'560\d{7}', line)
        found_orders.extend(matches)
        
        # Pattern très général: 10 chiffres commençant par 56
        matches = re.findall(r'56\d{8}', line)
        found_orders.extend(matches)
    
    # Déduplication et normalisation
    normalized_orders = []
    for order in found_orders:
        normalized = normalize_order_number(order)
        if normalized and normalized not in normalized_orders:
            normalized_orders.append(normalized)
    
    # Retourner le premier ordre trouvé (ou None si aucun)
    return normalized_orders[0] if normalized_orders else None

def extract_invoice_info(pdf_lines):
    invoice_id = ""
    net_total = None
    commande = None

    # Extraction du numéro de commande
    commande = extract_commande_from_pdf(pdf_lines)

    # Extraction du numéro de facture
    for i, line in enumerate(pdf_lines):
        norm_line = normalize(line)
        if "invoice id/number" in norm_line or "numéro" in norm_line:
            for j in range(i + 1, min(i + 3, len(pdf_lines))):
                next_line = pdf_lines[j].strip()
                if next_line and not "purchase" in normalize(next_line):
                    invoice_id = next_line
                    break
            break

    # Extraction du total net
    for line in pdf_lines:
        norm_line = normalize(line)
        if "invoice total" in norm_line:
            parts = line.replace(",", ".").split()
            for idx, p in enumerate(parts):
                if "total" in normalize(p):
                    try:
                        net_total = float(parts[idx + 1])
                    except:
                        pass
                    break
            break

    return invoice_id, net_total, commande

def match_row_to_pdf(row, pdf_lines, fuzzy=True, fuzzy_threshold=85):
    results = []
    cmd = normalize(str(row.get("N° commande", "")))
    montant = normalize(str(row.get("Montant brut", "")))
    taux = normalize(str(row.get("Taux de facturation", "")))
    rubrique = normalize(str(row.get("Code rubrique", "")))
    unite = normalize(str(row.get("Unités", "")))
    semaine = normalize(str(row.get("Semaine finissant le", "")))

    for line in pdf_lines:
        norm_line = normalize(line)
        if cmd in norm_line:
            score = 0
            matched = []
            if montant and (montant in norm_line or (fuzzy and fuzz.partial_ratio(montant, norm_line) > fuzzy_threshold)):
                score += 1
                matched.append("montant")
            if taux and (taux in norm_line or (fuzzy and fuzz.partial_ratio(taux, norm_line) > fuzzy_threshold)):
                score += 1
                matched.append("taux")
            if rubrique and (rubrique in norm_line or (fuzzy and fuzz.partial_ratio(rubrique, norm_line) > fuzzy_threshold)):
                score += 1
                matched.append("rubrique")
            if unite and (unite in norm_line or (fuzzy and fuzz.partial_ratio(unite, norm_line) > fuzzy_threshold)):
                score += 1
                matched.append("unités")
            if semaine and (semaine in norm_line or (fuzzy and fuzz.partial_ratio(semaine, norm_line) > fuzzy_threshold)):
                score += 1
                matched.append("semaine")
            results.append((score, line, matched))
    if results:
        results = sorted(results, key=lambda x: -x[0])
        return results[0][1], results[0][0], results[0][2]
    return "", 0, []

def generate_html_report(results, columns, synthese_factures, logo_url=None):
    template = jinja2.Template("""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <title>Rapport de rapprochement Excel ↔ PDF</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 2em;}
            h1 { color: #003366;}
            table { border-collapse: collapse; width: 100%; margin-top: 1em;}
            th, td { border: 1px solid #ddd; padding: 8px; }
            th { background-color: #f2f2f2; }
            tr:nth-child(even) { background-color: #f9f9f9;}
            .matched { background: #d4edda;}
            .unmatched { background: #f8d7da;}
            .score-high { color: #155724;}
            .score-low { color: #721c24;}
            .facture-ok { background: #d4edda; }
            .facture-ko { background: #f8d7da; }
        </style>
    </head>
    <body>
        {% if logo_url %}
        <img src="{{ logo_url }}" alt="Logo" style="height:60px;"/>
        {% endif %}
        <h1>Rapport de rapprochement Excel ↔ PDF</h1>
        <h2>Synthèse par facture</h2>
        <table>
            <tr>
                <th>Fichier Excel</th>
                <th>Fichier PDF</th>
                <th>Numéro de facture</th>
                <th>N° Commande PDF</th>
                <th>Total net PDF</th>
                <th>Total net Excel</th>
                <th>Statut</th>
            </tr>
            {% for syn in synthese_factures %}
            <tr class="{{ 'facture-ok' if syn['Statut'] == 'OK' else 'facture-ko' }}">
                <td>{{ syn['Fichier Excel'] }}</td>
                <td>{{ syn['Fichier PDF'] }}</td>
                <td>{{ syn['Numéro facture'] }}</td>
                <td>{{ syn['N° Commande PDF'] }}</td>
                <td>{{ "{:,.2f}".format(syn['Total net PDF']) if syn['Total net PDF'] is not none else '' }}</td>
                <td>{{ "{:,.2f}".format(syn['Total net Excel']) if syn['Total net Excel'] is not none else '' }}</td>
                <td>{{ syn['Statut'] }}</td>
            </tr>
            {% endfor %}
        </table>
        <h2>Rapprochement ligne à ligne</h2>
        <table>
            <tr>
                {% for col in columns %}
                <th>{{ col }}</th>
                {% endfor %}
                <th>Fichier Excel</th>
                <th>PDF correspondant</th>
                <th>Ligne PDF</th>
                <th>Score</th>
                <th>Champs trouvés</th>
            </tr>
            {% for row in results %}
            <tr class="{{ 'matched' if row['Score correspondance'] > 0 else 'unmatched' }}">
                {% for col in columns %}
                <td>{{ row[col] }}</td>
                {% endfor %}
                <td>{{ row['Fichier Excel'] }}</td>
                <td>{{ row['PDF correspondant'] }}</td>
                <td>{{ row['Ligne PDF correspondante'] }}</td>
                <td class="{{ 'score-high' if row['Score correspondance'] > 2 else 'score-low' }}">{{ row['Score correspondance'] }}</td>
                <td>{{ row['Champs trouvés'] }}</td>
            </tr>
            {% endfor %}
        </table>
        <p style="margin-top:2em;font-size:small;color:#888;">Généré le {{ now }}</p>
    </body>
    </html>
    """)
    return template.render(results=results, columns=columns, synthese_factures=synthese_factures, logo_url=logo_url, now=datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))

def generate_excel_report(df, excel_columns, synthese_factures):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapprochement"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="003366")
    for col_num, col in enumerate(excel_columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(40, max_length + 2))
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Synthèse Factures")
    syn_cols = ["Fichier Excel", "Fichier PDF", "Numéro facture", "N° Commande PDF", "Total net PDF", "Total net Excel", "Statut"]
    for col_num, col in enumerate(syn_cols, 1):
        cell = ws2.cell(row=1, column=col_num, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, syn in enumerate(synthese_factures, 2):
        ws2.cell(row=row_idx, column=1, value=syn["Fichier Excel"])
        ws2.cell(row=row_idx, column=2, value=syn["Fichier PDF"])
        ws2.cell(row=row_idx, column=3, value=syn["Numéro facture"])
        ws2.cell(row=row_idx, column=4, value=syn["N° Commande PDF"])
        ws2.cell(row=row_idx, column=5, value=syn["Total net PDF"])
        ws2.cell(row=row_idx, column=6, value=syn["Total net Excel"])
        ws2.cell(row=row_idx, column=7, value=syn["Statut"])
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(12, min(40, max_length + 2))
    ws2.auto_filter.ref = ws2.dimensions

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

st.title("Rapprochement automatique Excel Beeline ↔ PDF par N° de commande")
st.info("Cette version améliore le matching en recherchant automatiquement les N° de commande dans les PDFs")

pdf_files = st.file_uploader("Importer un ou plusieurs PDF", type="pdf", accept_multiple_files=True)
excel_files = st.file_uploader("Importer un ou plusieurs fichiers Excel Beeline", type=["xlsx", "xls"], accept_multiple_files=True)
fuzzy = st.checkbox("Tolérance de rapprochement (fuzzy)", value=True)
fuzzy_threshold = st.slider("Seuil fuzzy (plus élevé = plus strict)", 70, 100, 85) if fuzzy else 100
debug_mode = st.checkbox("🔍 Mode debug - Afficher les numéros de commande extraits", value=False)

if excel_files and pdf_files:
    st.info("Traitement en cours...")

    columns = ["Collaborateur", "N° commande", "Montant brut", "Montant net à payer au fournisseur", "Montant de la taxe",
               "Taux de facturation", "Code rubrique", "Unités", "Semaine finissant le"]
    excel_columns = columns + ["Fichier Excel", "PDF correspondant", "Ligne PDF correspondante", "Score correspondance", "Champs trouvés"]

    # Étape 1: Parser tous les PDFs et extraire les N° de commande
    st.subheader("Étape 1: Analyse des PDFs")
    parsed_pdfs = []
    pdf_infos = {}
    pdf_by_commande = {}  # Nouveau: indexer les PDFs par numéro de commande
    all_pdf_orders = []  # Pour le debug
    
    for pdf_file in pdf_files:
        lines = extract_pdf_lines(pdf_file)
        invoice_id, net_total, commande = extract_invoice_info(lines)
        
        # Normaliser le numéro de commande
        commande_normalized = normalize_order_number(commande) if commande else None
        
        parsed_pdfs.append({'filename': pdf_file.name, 'lines': lines})
        pdf_infos[pdf_file.name] = {
            "Numéro facture": invoice_id,
            "Total net PDF": net_total,
            "N° Commande": commande_normalized
        }
        
        # Ajouter à la liste debug
        if commande_normalized:
            all_pdf_orders.append(commande_normalized)
        
        # Indexer par numéro de commande
        if commande_normalized:
            if commande_normalized not in pdf_by_commande:
                pdf_by_commande[commande_normalized] = []
            pdf_by_commande[commande_normalized].append({
                'filename': pdf_file.name,
                'lines': lines,
                'invoice_id': invoice_id,
                'net_total': net_total
            })
            st.success(f"PDF: {pdf_file.name} → Commande: {commande_normalized}")
        else:
            st.warning(f"PDF: {pdf_file.name} → Commande non trouvée")

    # Debug: Affichage des numéros extraits des PDFs
    if debug_mode and all_pdf_orders:
        st.info("🔍 **DEBUG - Numéros de commande extraits des PDFs:**")
        st.write(f"Total: {len(all_pdf_orders)} numéros | Uniques: {len(set(all_pdf_orders))}")
        st.code(f"Numéros PDF: {sorted(set(all_pdf_orders))}")

    st.subheader("Étape 2: Matching intelligent par N° de commande")
    
    # Étape 2: Grouper les lignes Excel par N° de commande avec normalisation
    excel_by_commande = {}
    all_excel_rows = []
    all_excel_orders = []  # Pour le debug
    
    for excel_file in excel_files:
        df = pd.read_excel(excel_file)
        for idx, row in df.iterrows():
            # Normaliser le numéro de commande Excel
            commande_raw = row.get("N° commande", "")
            commande_normalized = normalize_order_number(commande_raw)
            
            row_data = {
                'excel_file': excel_file.name,
                'row': row,
                'commande': commande_normalized,
                'idx': idx
            }
            all_excel_rows.append(row_data)
            
            # Ajouter à la liste debug
            if commande_normalized:
                all_excel_orders.append(commande_normalized)
            
            if commande_normalized:
                if commande_normalized not in excel_by_commande:
                    excel_by_commande[commande_normalized] = []
                excel_by_commande[commande_normalized].append(row_data)

    # Debug: Affichage des numéros extraits d'Excel
    if debug_mode and all_excel_orders:
        st.info("🔍 **DEBUG - Numéros de commande extraits d'Excel:**")
        st.write(f"Total: {len(all_excel_orders)} numéros | Uniques: {len(set(all_excel_orders))}")
        st.code(f"Numéros Excel: {sorted(set(all_excel_orders))}")
        
        # Comparaison des listes
        pdf_set = set(all_pdf_orders)
        excel_set = set(all_excel_orders)
        common = pdf_set & excel_set
        pdf_only = pdf_set - excel_set
        excel_only = excel_set - pdf_set
        
        st.info("🔍 **DEBUG - Comparaison des numéros:**")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Correspondances", len(common))
            if common:
                st.success(f"Communs: {sorted(common)}")
        with col2:
            st.metric("PDF uniquement", len(pdf_only))
            if pdf_only:
                st.warning(f"PDF seulement: {sorted(pdf_only)}")
        with col3:
            st.metric("Excel uniquement", len(excel_only))
            if excel_only:
                st.warning(f"Excel seulement: {sorted(excel_only)}")

    # Étape 3: Matching intelligent
    synthese_factures = []
    results = []
    matched_pdfs = set()
    matched_excel = set()

    # Matcher par numéro de commande
    for commande, excel_rows in excel_by_commande.items():
        if commande in pdf_by_commande:
            pdf_matches = pdf_by_commande[commande]
            
            for pdf_match in pdf_matches:
                # Calculer le total Excel pour cette commande
                total_excel = sum(float(r['row'].get("Montant net à payer au fournisseur", 0)) 
                                for r in excel_rows)
                
                # Valider le total
                total_pdf = pdf_match['net_total']
                statut = "OK" if (total_pdf and abs(total_pdf - total_excel) < 0.02) else "Écart détecté"
                
                synthese_factures.append({
                    "Fichier Excel": ", ".join(set(r['excel_file'] for r in excel_rows)),
                    "Fichier PDF": pdf_match['filename'],
                    "Numéro facture": pdf_match['invoice_id'],
                    "N° Commande PDF": commande,
                    "Total net PDF": total_pdf,
                    "Total net Excel": total_excel,
                    "Statut": statut
                })
                
                matched_pdfs.add(pdf_match['filename'])
                
                # Matcher chaque ligne Excel avec le PDF
                for excel_row in excel_rows:
                    matched_line, score, matched_fields = match_row_to_pdf(
                        excel_row['row'], 
                        pdf_match['lines'], 
                        fuzzy, 
                        fuzzy_threshold
                    )
                    
                    results.append({
                        **{col: excel_row['row'].get(col, "") for col in columns},
                        "Fichier Excel": excel_row['excel_file'],
                        "PDF correspondant": pdf_match['filename'],
                        "Ligne PDF correspondante": matched_line,
                        "Score correspondance": score,
                        "Champs trouvés": ', '.join(matched_fields)
                    })
                    
                    matched_excel.add(f"{excel_row['excel_file']}_{idx}")

    # Ajouter les non-matchés
    for pdf_file in pdf_files:
        if pdf_file.name not in matched_pdfs:
            pdf_info = pdf_infos.get(pdf_file.name, {})
            synthese_factures.append({
                "Fichier Excel": "",
                "Fichier PDF": pdf_file.name,
                "Numéro facture": pdf_info.get("Numéro facture", ""),
                "N° Commande PDF": pdf_info.get("N° Commande", ""),
                "Total net PDF": pdf_info.get("Total net PDF"),
                "Total net Excel": None,
                "Statut": "PDF sans Excel"
            })

    for excel_row_data in all_excel_rows:
        key = f"{excel_row_data['excel_file']}_{excel_row_data.get('idx', '')}"
        if key not in matched_excel:
            results.append({
                **{col: excel_row_data['row'].get(col, "") for col in columns},
                "Fichier Excel": excel_row_data['excel_file'],
                "PDF correspondant": "",
                "Ligne PDF correspondante": "",
                "Score correspondance": 0,
                "Champs trouvés": ""
            })

    res_df = pd.DataFrame(results)
    res_df = res_df[excel_columns]

    st.subheader("Résultats du rapprochement")
    
    # Afficher les statistiques
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("PDFs traités", len(pdf_files))
    with col2:
        st.metric("PDFs matchés", len(matched_pdfs))
    with col3:
        st.metric("Commandes communes", len(set(excel_by_commande.keys()) & set(pdf_by_commande.keys())))
    with col4:
        st.metric("Lignes matchées", len([r for r in results if r['Score correspondance'] > 0]))
    
    st.dataframe(res_df, use_container_width=True)

    with st.expander("Filtrer les résultats"):
        col1, col2 = st.columns(2)
        with col1:
            search_collab = st.text_input("Recherche par nom d'intérimaire")
            min_score = st.slider("Score minimum", 0, 5, 0)
        with col2:
            only_matched = st.checkbox("Afficher seulement les correspondances")
        filtered = res_df
        if search_collab:
            filtered = filtered[filtered["Collaborateur"].str.contains(search_collab, case=False, na=False)]
        filtered = filtered[filtered["Score correspondance"] >= min_score]
        if only_matched:
            filtered = filtered[filtered["Score correspondance"] > 0]
        st.dataframe(filtered, use_container_width=True)

    html_report = generate_html_report(filtered.to_dict(orient="records"), columns, synthese_factures)
    st.subheader("Télécharger le rapport HTML")
    st.download_button(
        label="Télécharger le rapport HTML",
        data=html_report,
        file_name="rapport_rapprochement.html",
        mime="text/html"
    )

    st.subheader("Télécharger le rapport Excel")
    excel_file_bytes = generate_excel_report(filtered, excel_columns, synthese_factures)
    st.download_button(
        label="Télécharger le rapport Excel",
        data=excel_file_bytes,
        file_name="rapport_rapprochement.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
